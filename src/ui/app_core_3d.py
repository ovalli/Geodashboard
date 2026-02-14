# ======================================================
# src/ui/app_core_3d.py  (COMPLET)
# ✅ Onglets Coupes en 3D avec indicateur couleur (emoji) basé sur Coupe.color
# ✅ Chantier complet : points colorés selon la coupe qui contient la cible
#     (fallback "red" si cible non assignée à une coupe)
#
# ⚠️ Ne modifie pas le design / comportement général :
# - Même UI 3D, mêmes options, mêmes sliders
# - Juste : onglets + couleurs de points en chantier complet
# - ✅ MODIF: lecture Mesures Completes = 1ère feuille (peu importe le nom)
#
# ✅ MODIF (tech): Delaunay SciPy redevient prioritaire (surfaces propres),
#    fallback alphahull uniquement si SciPy indisponible.
# ======================================================

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Tuple

import numpy as np
import streamlit as st
from streamlit_option_menu import option_menu

from src.io.coupes_manager import CoupesManager


# ======================================================
# Constants (conservatifs)
# ======================================================
HEAD_REL = 0.055
LINE_REL = 0.015
CONE_REL = 0.090

HEAD_ABS_MIN = 1.4
LINE_ABS_MIN = 0.5
CONE_ABS_MIN = 1.8

VECTOR_LINE_WIDTH = 6

HITBOX_SIZE = 24
POINT_SIZE = 7


# ======================================================
# Palette & emoji mapping (comme Coups depuis CST)
# ======================================================
_ZONE_PALETTE = ["#146EFF", "#22C55E", "#F97316", "#A855F7", "#EF4444", "#06B6D4", "#EAB308", "#EC4899"]
_ZONE_EMOJI = ["🔵", "🟢", "🟠", "🟣", "🔴", "🟦", "🟡", "🩷"]

_COLOR_TO_EMOJI = {
    "#146EFF": "🔵",
    "#22C55E": "🟢",
    "#F97316": "🟠",
    "#A855F7": "🟣",
    "#EF4444": "🔴",
    "#06B6D4": "🟦",
    "#EAB308": "🟡",
    "#EC4899": "🩷",
}


def _zone_color(i: int) -> str:
    return _ZONE_PALETTE[i % len(_ZONE_PALETTE)]


def _zone_emoji(i: int) -> str:
    return _ZONE_EMOJI[i % len(_ZONE_EMOJI)]


def _emoji_from_color(col: str, fallback_i: int) -> str:
    col = (col or "").strip()
    return _COLOR_TO_EMOJI.get(col, _zone_emoji(fallback_i))


# ======================================================
# Paths
# ======================================================
def _project_root() -> Path:
    here = Path(__file__).resolve()
    for p in [here.parent, *here.parents]:
        if (p / "data" / "common_data").exists():
            return p
        if (p / "app.py").exists():
            return p
    return here.parents[2]


def _find_mesures_completes_xlsx(common_data_dir: Path) -> Path | None:
    if not common_data_dir.exists():
        return None
    needles = ["mesures completes", "mesures complètes", "mesures complete", "mesures compl"]
    candidates: list[Path] = []
    for p in common_data_dir.iterdir():
        if p.is_file() and p.suffix.lower() == ".xlsx":
            stem = p.stem.lower()
            if any(n in stem for n in needles):
                candidates.append(p)
    if not candidates:
        return None
    candidates.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return candidates[0]


# ======================================================
# Utilities
# ======================================================
def _nice_scales(min_pow: int = 0, max_pow: int = 6) -> List[int]:
    vals: list[int] = []
    for p in range(min_pow, max_pow + 1):
        base = 10**p
        vals.extend([1 * base, 2 * base, 5 * base])
    vals.append(10 ** (max_pow + 1))
    return sorted(set(vals))


@st.cache_data(show_spinner=False)
def _first_sheet_name(mesures_xlsx: str) -> str:
    """
    ✅ Toujours lire la 1ère feuille du classeur Mesures Completes,
    quel que soit son nom (Data, data, Données, etc.)
    """
    import openpyxl

    wb = openpyxl.load_workbook(mesures_xlsx, data_only=True, read_only=True)
    if not wb.sheetnames:
        raise ValueError("Classeur Mesures Completes sans onglets.")
    return wb.sheetnames[0]


@st.cache_data(show_spinner=False)
def _build_target_col_map(mesures_xlsx: str, sheet_name: str) -> Dict[str, int]:
    import openpyxl

    wb = openpyxl.load_workbook(mesures_xlsx, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    out: Dict[str, int] = {}
    for idx, v in enumerate(row1, start=1):
        if isinstance(v, str) and v.strip():
            out[v.strip()] = idx
    return out


@st.cache_data(show_spinner=False)
def _list_all_targets_from_mesures_header(mesures_xlsx: str, sheet_name: str) -> List[str]:
    col_map = _build_target_col_map(mesures_xlsx, sheet_name)
    items = sorted(col_map.items(), key=lambda kv: kv[1])

    targets: list[str] = []
    for name, col in items:
        if col == 1 and isinstance(name, str) and name.lower() in {"date", "dates", "time", "temps"}:
            continue
        targets.append(name)

    seen = set()
    out: list[str] = []
    for t in targets:
        if t not in seen:
            seen.add(t)
            out.append(t)
    return out


@st.cache_data(show_spinner=False)
def _first_and_last_coords_for_targets(
    mesures_xlsx: str,
    sheet_name: str,
    targets: Tuple[str, ...],
) -> Tuple[str, str, Dict[str, Tuple[float, float, float]], Dict[str, Tuple[float, float, float]]]:
    import openpyxl

    col_map = _build_target_col_map(mesures_xlsx, sheet_name)
    triplets: Dict[str, Tuple[int, int, int]] = {}
    for t in targets:
        c = col_map.get(t)
        if c is not None:
            triplets[t] = (c, c + 1, c + 2)

    wb = openpyxl.load_workbook(mesures_xlsx, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    first_date = None
    last_date = None
    coords_first: Dict[str, Tuple[float, float, float]] = {}
    coords_last: Dict[str, Tuple[float, float, float]] = {}

    def _fmt_date(d) -> str:
        if d is None:
            return "date inconnue"
        if hasattr(d, "strftime"):
            return d.strftime("%Y-%m-%d")
        return str(d)

    for row in ws.iter_rows(min_row=2, values_only=True):
        d = row[0]
        if d is not None:
            if first_date is None:
                first_date = d
            last_date = d

        for t, (cx, cy, cz) in triplets.items():
            x = row[cx - 1] if cx - 1 < len(row) else None
            y = row[cy - 1] if cy - 1 < len(row) else None
            z = row[cz - 1] if cz - 1 < len(row) else None
            if x is None or y is None or z is None:
                continue
            try:
                pt = (float(x), float(y), float(z))
            except Exception:
                continue

            if t not in coords_first:
                coords_first[t] = pt
            coords_last[t] = pt

    return _fmt_date(first_date), _fmt_date(last_date), coords_first, coords_last


def _triangles_from_xy(x: np.ndarray, y: np.ndarray) -> Tuple[np.ndarray, np.ndarray, np.ndarray, bool]:
    """
    Retourne (i,j,k, used_scipy).
    - Si SciPy dispo => Delaunay => surface propre
    - Sinon => arrays vides + used_scipy=False (fallback alphahull)
    """
    if len(x) < 3:
        return np.array([], dtype=int), np.array([], dtype=int), np.array([], dtype=int), False

    try:
        from scipy.spatial import Delaunay  # type: ignore
    except Exception:
        return np.array([], dtype=int), np.array([], dtype=int), np.array([], dtype=int), False

    pts = np.column_stack((x, y))
    try:
        tri = Delaunay(pts)
        simp = tri.simplices
        if simp is None or len(simp) == 0:
            return np.array([], dtype=int), np.array([], dtype=int), np.array([], dtype=int), False
        return simp[:, 0], simp[:, 1], simp[:, 2], True
    except Exception:
        return np.array([], dtype=int), np.array([], dtype=int), np.array([], dtype=int), False


def _robust_keep_mask(x: np.ndarray, y: np.ndarray, z: np.ndarray, k: float) -> np.ndarray:
    pts = np.column_stack([x, y, z])
    med = np.median(pts, axis=0)
    d = np.linalg.norm(pts - med, axis=1)

    d_med = float(np.median(d))
    mad = float(np.median(np.abs(d - d_med)))

    if mad > 1e-12:
        rz = 0.6745 * (d - d_med) / mad
        return rz <= k

    q1 = float(np.percentile(d, 25))
    q3 = float(np.percentile(d, 75))
    iqr = q3 - q1
    if iqr <= 1e-12:
        return np.ones(len(d), dtype=bool)

    factor = max(1.5, min(6.0, k / 3.0))
    thr = q3 + factor * iqr
    return d <= thr


def _render_pretty_list(items: List[str], cols: int = 3) -> None:
    items = sorted([str(x) for x in items if str(x).strip()])
    if not items:
        st.write("—")
        return

    n = len(items)
    cols = max(1, int(cols))
    rows = int(np.ceil(n / cols))
    grid = [items[i * rows : (i + 1) * rows] for i in range(cols)]
    c = st.columns(cols)
    for j in range(cols):
        with c[j]:
            for it in grid[j]:
                st.write(f"- {it}")


def _add_xyz_triad(fig, xs, ys, zs) -> None:
    import plotly.graph_objects as go

    xmin, xmax = float(xs.min()), float(xs.max())
    ymin, ymax = float(ys.min()), float(ys.max())
    zmin, zmax = float(zs.min()), float(zs.max())

    dx = max(xmax - xmin, 1.0)
    dy = max(ymax - ymin, 1.0)
    dz = max(zmax - zmin, 1.0)

    diag = float(np.sqrt(dx * dx + dy * dy + dz * dz))
    L = max(diag * 0.08, 1.0)

    ox = xmin + dx * 0.06
    oy = ymin + dy * 0.06
    oz = zmin + dz * 0.06

    fig.add_trace(
        go.Scatter3d(
            x=[ox, ox + L, None, ox, ox, None, ox, ox, None],
            y=[oy, oy, None, oy, oy + L, None, oy, oy, None],
            z=[oz, oz, None, oz, oz, None, oz, oz + L, None],
            mode="lines",
            hoverinfo="skip",
            line=dict(width=6, color="rgba(190,190,190,0.85)"),
            showlegend=False,
            name="",
        )
    )
    fig.add_trace(
        go.Scatter3d(
            x=[ox + L, ox, ox],
            y=[oy, oy + L, oy],
            z=[oz, oz, oz + L],
            mode="text",
            text=["X=Est", "Y=Nord", "Z"],
            hoverinfo="skip",
            textfont=dict(size=14, color="rgba(190,190,190,0.95)"),
            showlegend=False,
            name="",
        )
    )


# ======================================================
# Core plot
# ======================================================
def _plot_3d(
    title: str,
    selected_key: str,
    targets: List[str],
    mesures_path: Path,
    show_surface: bool,
    show_vectors: bool,
    scale: float,
    *,
    auto_filter: bool,
    outlier_k: float,
    motion_mm_range: Tuple[float, float],
    target_color_map: Dict[str, str] | None = None,
    default_point_color: str = "red",
) -> None:
    import plotly.graph_objects as go

    sheet = _first_sheet_name(str(mesures_path))

    date_first, date_last, coords_first, coords_last = _first_and_last_coords_for_targets(
        str(mesures_path), sheet, tuple(targets)
    )

    points_last: List[Tuple[str, float, float, float]] = []
    missing: List[str] = []
    for t in targets:
        if t in coords_last:
            x, y, z = coords_last[t]
            points_last.append((t, x, y, z))
        else:
            missing.append(t)

    if not points_last:
        st.warning("Aucune coordonnée XYZ exploitable.")
        return

    names_all = [p[0] for p in points_last]
    xs_all = np.array([p[1] for p in points_last], dtype=float)
    ys_all = np.array([p[2] for p in points_last], dtype=float)
    zs_all = np.array([p[3] for p in points_last], dtype=float)

    removed_outliers: List[str] = []
    if auto_filter and len(xs_all) >= 5:
        mask = _robust_keep_mask(xs_all, ys_all, zs_all, float(outlier_k))
        if mask.sum() >= 3:
            removed_outliers = [names_all[i] for i in range(len(names_all)) if not mask[i]]
            names = [names_all[i] for i in range(len(names_all)) if mask[i]]
            xs = xs_all[mask]
            ys = ys_all[mask]
            zs = zs_all[mask]
        else:
            names = names_all
            xs, ys, zs = xs_all, ys_all, zs_all
            removed_outliers = []
    else:
        names = names_all
        xs, ys, zs = xs_all, ys_all, zs_all

    xmin, xmax = float(xs.min()), float(xs.max())
    ymin, ymax = float(ys.min()), float(ys.max())
    zmin, zmax = float(zs.min()), float(zs.max())
    diag = float(np.sqrt((xmax - xmin) ** 2 + (ymax - ymin) ** 2 + (zmax - zmin) ** 2))
    diag = max(diag, 1.0)

    HEAD_LEN = max(diag * float(HEAD_REL), float(HEAD_ABS_MIN))
    LINE_MIN = max(diag * float(LINE_REL), float(LINE_ABS_MIN))
    CONE_SIZEREF = max(diag * float(CONE_REL), float(CONE_ABS_MIN))

    min_mm, max_mm = float(motion_mm_range[0]), float(motion_mm_range[1])
    min_mm = max(0.0, min_mm)
    max_mm = max(min_mm, max_mm)

    st.caption(
        f"{title} · Mesures: **{mesures_path.name}** · Feuille: **{sheet}** · Première: **{date_first}** · Dernière: **{date_last}** · "
        f"XYZ: **{len(points_last)} / {len(targets)}** · Outliers exclus: **{len(removed_outliers)}** · k={outlier_k:g}"
        + (f" · Vecteurs: mm ∈ [{min_mm:g}, {max_mm:g}]" if show_vectors else "")
    )

    if removed_outliers:
        with st.expander(f"Cibles exclues (outliers) ({len(removed_outliers)})"):
            _render_pretty_list(removed_outliers, cols=3)

    keep_set = set(names)
    i, j, k, used_scipy = _triangles_from_xy(xs, ys)

    # warning une seule fois si fallback
    if show_surface and (not used_scipy) and ("warned_no_scipy_delaunay" not in st.session_state):
        st.session_state["warned_no_scipy_delaunay"] = True
        st.warning("⚠️ SciPy non disponible : surface calculée en mode fallback (qualité moindre). Installe `scipy` pour retrouver la surface propre.")

    vline_x: list[float] = []
    vline_y: list[float] = []
    vline_z: list[float] = []

    cone_x: list[float] = []
    cone_y: list[float] = []
    cone_z: list[float] = []
    cone_u: list[float] = []
    cone_v: list[float] = []
    cone_w: list[float] = []

    if show_vectors:
        for t in targets:
            if t not in keep_set:
                continue
            if t not in coords_first or t not in coords_last:
                continue

            xF, yF, zF = coords_first[t]
            xL, yL, zL = coords_last[t]

            dx = xL - xF
            dy = yL - yF
            dz = zL - zF

            mag_m = float(np.sqrt(dx * dx + dy * dy + dz * dz))
            if mag_m <= 1e-12:
                continue

            motion_mm = mag_m * 1000.0
            if motion_mm < min_mm or motion_mm > max_mm:
                continue

            ux = dx / mag_m
            uy = dy / mag_m
            uz = dz / mag_m

            disp_len = mag_m / float(scale)
            total_len = max(disp_len, float(HEAD_LEN) + float(LINE_MIN))

            tip_x = xL + ux * total_len
            tip_y = yL + uy * total_len
            tip_z = zL + uz * total_len

            base_x = tip_x - ux * float(HEAD_LEN)
            base_y = tip_y - uy * float(HEAD_LEN)
            base_z = tip_z - uz * float(HEAD_LEN)

            vline_x += [xL, base_x, None]
            vline_y += [yL, base_y, None]
            vline_z += [zL, base_z, None]

            cone_x.append(base_x)
            cone_y.append(base_y)
            cone_z.append(base_z)
            cone_u.append(ux * float(HEAD_LEN))
            cone_v.append(uy * float(HEAD_LEN))
            cone_w.append(uz * float(HEAD_LEN))

    # ✅ point colors
    if target_color_map is None:
        point_colors = default_point_color
    else:
        point_colors = [str(target_color_map.get(n, default_point_color) or default_point_color) for n in names]

    fig = go.Figure()
    hover_tpl = "<b>%{customdata}</b><br>Z=%{z:.1f}<extra></extra>"
    custom = np.array(names, dtype=object)

    # invisible hitbox
    fig.add_trace(
        go.Scatter3d(
            x=xs,
            y=ys,
            z=zs,
            mode="markers",
            marker=dict(size=HITBOX_SIZE, opacity=0.0, color="rgba(255,0,0,0)"),
            customdata=custom,
            hovertemplate=hover_tpl,
            showlegend=False,
            name="",
        )
    )

    # visible points
    fig.add_trace(
        go.Scatter3d(
            x=xs,
            y=ys,
            z=zs,
            mode="markers",
            marker=dict(size=POINT_SIZE, opacity=1.0, color=point_colors),
            customdata=custom,
            hovertemplate=hover_tpl,
            showlegend=False,
            name="",
        )
    )

    # surface (Delaunay prioritaire, fallback alphahull sinon)
    if used_scipy and len(i) > 0:
        fig.add_trace(
            go.Mesh3d(
                x=xs,
                y=ys,
                z=zs,
                i=i,
                j=j,
                k=k,
                intensity=zs,
                colorscale=[[0.0, "saddlebrown"], [1.0, "seagreen"]],
                showscale=False,
                opacity=0.95,
                flatshading=True,
                hoverinfo="skip",
                showlegend=False,
                visible=bool(show_surface and len(i) > 0),
                name="",
            )
        )
    else:
        fig.add_trace(
            go.Mesh3d(
                x=xs,
                y=ys,
                z=zs,
                alphahull=0,
                intensity=zs,
                colorscale=[[0.0, "saddlebrown"], [1.0, "seagreen"]],
                showscale=False,
                opacity=0.95,
                flatshading=True,
                hoverinfo="skip",
                showlegend=False,
                visible=bool(show_surface),
                name="",
            )
        )

    # vectors line + cone
    if show_vectors and len(vline_x) > 0:
        fig.add_trace(
            go.Scatter3d(
                x=vline_x,
                y=vline_y,
                z=vline_z,
                mode="lines",
                hoverinfo="skip",
                line=dict(color="orange", width=int(VECTOR_LINE_WIDTH)),
                showlegend=False,
                name="",
            )
        )

    if show_vectors and len(cone_x) > 0:
        fig.add_trace(
            go.Cone(
                x=cone_x,
                y=cone_y,
                z=cone_z,
                u=cone_u,
                v=cone_v,
                w=cone_w,
                anchor="tail",
                cauto=False,
                cmin=0.0,
                cmax=1.0,
                colorscale=[[0.0, "orange"], [1.0, "orange"]],
                showscale=False,
                sizemode="absolute",
                sizeref=float(CONE_SIZEREF),
                opacity=0.98,
                hoverinfo="skip",
                name="",
                lighting=dict(
                    ambient=1.0,
                    diffuse=0.0,
                    specular=0.0,
                    roughness=1.0,
                    fresnel=0.0,
                ),
            )
        )

    _add_xyz_triad(fig, xs, ys, zs)

    fig.update_layout(
        height=720,
        margin=dict(l=0, r=0, t=0, b=0),
        showlegend=False,
        scene=dict(
            xaxis=dict(visible=False),
            yaxis=dict(visible=False),
            zaxis=dict(visible=False),
            bgcolor="rgba(0,0,0,0)",
            aspectmode="data",
            dragmode="orbit",
        ),
        uirevision=f"3d::{selected_key}",
        hovermode="closest",
    )

    st.plotly_chart(
        fig,
        use_container_width=True,
        key=f"plot3d::{selected_key}",
        config={"displaylogo": False, "scrollZoom": True},
    )

    if missing:
        with st.expander(f"Sans XYZ trouvés ({len(missing)})"):
            _render_pretty_list(missing, cols=3)


# ======================================================
# Public UI
# ======================================================
def render_3d() -> None:
    root = _project_root()
    common_data = root / "data" / "common_data"

    mgr = CoupesManager()
    coupes = mgr.list_coupes()
    if not coupes:
        st.info("Aucune coupe détectée dans le JSON.")
        return

    mesures_path = _find_mesures_completes_xlsx(common_data)
    if mesures_path is None or not mesures_path.exists():
        st.error(f"Fichier Mesures Completes introuvable dans : {common_data}")
        return

    tab_coupe, tab_chantier = st.tabs(["Coupe", "Chantier Complet"])

    # ==================================================
    # Build maps:
    # - coupe_name -> color
    # - coupe_name -> emoji label for tabs
    # - target -> color (chantier complet)
    # ==================================================
    coupe_color_by_name: Dict[str, str] = {}
    coupe_label_by_name: Dict[str, str] = {}
    target_to_color: Dict[str, str] = {}
    duplicates: Dict[str, List[str]] = {}

    for i, c in enumerate(coupes):
        col = (getattr(c, "color", "") or "").strip() or _zone_color(i)
        coupe_color_by_name[c.name] = col
        coupe_label_by_name[c.name] = f"{_emoji_from_color(col, i)} {c.name}"

        for t in (c.targets or []):
            tt = str(t).strip()
            if not tt:
                continue
            if tt in target_to_color and target_to_color[tt] != col:
                duplicates.setdefault(tt, []).extend([target_to_color[tt], col])
                continue
            target_to_color[tt] = col

    # ==================================================
    # COUPE (avec onglets colorés)
    # ==================================================
    with tab_coupe:
        coupe_names = [c.name for c in coupes]
        display_options = [coupe_label_by_name[n] for n in coupe_names]
        display_to_name = {coupe_label_by_name[n]: n for n in coupe_names}

        if "cst_selected_coupe_3d" not in st.session_state:
            st.session_state["cst_selected_coupe_3d"] = coupe_names[0]

        default_name = st.session_state["cst_selected_coupe_3d"]
        if default_name not in coupe_names:
            default_name = coupe_names[0]
        default_disp = coupe_label_by_name[default_name]

        selected_disp = option_menu(
            menu_title=None,
            options=display_options,
            icons=[""] * len(display_options),
            orientation="horizontal",
            key="cst_coupe_tabs_3d",
            default_index=display_options.index(default_disp) if default_disp in display_options else 0,
            styles={
                "container": {"padding": "0!important"},
                "icon": {"display": "none"},
                "nav-link": {
                    "font-size": "0.90rem",
                    "padding": "8px 10px",
                    "margin": "0 4px 0 0",
                    "border-radius": "0.6rem",
                    "white-space": "nowrap",
                },
                "nav-link-hover": {"background-color": "rgba(0,0,0,0.06)"},
                "nav-link-selected": {"background-color": "rgba(0,0,0,0.10)", "font-weight": "700"},
            },
        )

        selected_name = display_to_name.get(selected_disp, coupe_names[0])
        st.session_state["cst_selected_coupe_3d"] = selected_name

        coupe = next((c for c in coupes if c.name == selected_name), None)
        if coupe is None:
            st.error("Coupe introuvable.")
            return

        targets = [t for t in (coupe.targets or []) if isinstance(t, str) and t.strip()]
        st.markdown(f"### {selected_name}")
        if not targets:
            st.info("Cette coupe ne contient aucune cible.")
            return

        c1, c2, c3 = st.columns([1, 1, 2], vertical_alignment="center")
        with c1:
            show_surface = st.checkbox("Surfaces", value=True, key="show_surface_3d_coupe")
        with c2:
            show_vectors = st.checkbox("Mouvements (vecteurs)", value=False, key="show_vectors_3d_coupe")
        with c3:
            scale = 500.0
            if show_vectors:
                scales = _nice_scales(0, 6)
                default_val = 500 if 500 in scales else scales[0]
                if "vec_scale_3d_coupe" not in st.session_state:
                    st.session_state["vec_scale_3d_coupe"] = default_val
                scale = float(
                    st.select_slider(
                        "Échelle vecteurs",
                        options=scales,
                        value=st.session_state["vec_scale_3d_coupe"],
                        key="vec_scale_3d_coupe",
                    )
                )

        motion_minmax = (2.0, 100.0)
        if show_vectors:
            motion_minmax = st.slider(
                "Afficher seulement si mouvement (mm) entre :",
                min_value=0.0,
                max_value=2000.0,
                value=(2.0, 100.0),
                step=1.0,
                key="motion_mm_range_coupe",
            )

        st.markdown("#### Filtre points aberrants")
        f1, f2 = st.columns([1, 2], vertical_alignment="center")
        with f1:
            auto_filter = st.checkbox("Auto (outliers)", value=True, key="auto_filter_coupe")
        with f2:
            outlier_k = st.slider("Tolérance", 2.0, 20.0, 4.0, 0.5, key="outlier_k_coupe")

        coupe_color = coupe_color_by_name.get(selected_name, "red")
        _plot_3d(
            title="Coupe",
            selected_key=f"coupe::{selected_name}",
            targets=targets,
            mesures_path=mesures_path,
            show_surface=show_surface,
            show_vectors=show_vectors,
            scale=scale,
            auto_filter=auto_filter,
            outlier_k=float(outlier_k),
            motion_mm_range=(float(motion_minmax[0]), float(motion_minmax[1])),
            target_color_map={str(t): coupe_color for t in targets},
            default_point_color="red",
        )

    # ==================================================
    # CHANTIER COMPLET (points colorés par coupe)
    # ==================================================
    with tab_chantier:
        st.markdown("### Chantier Complet")

        if duplicates:
            with st.expander(f"Cibles présentes dans plusieurs coupes ({len(duplicates)})"):
                for t, cols in sorted(duplicates.items(), key=lambda kv: kv[0].lower()):
                    uniq = []
                    for c in cols:
                        if c not in uniq:
                            uniq.append(c)
                    st.write(f"- **{t}** : {', '.join(uniq)}")

        sheet = _first_sheet_name(str(mesures_path))
        all_targets = _list_all_targets_from_mesures_header(str(mesures_path), sheet)
        if not all_targets:
            st.error(f"Aucune cible détectée dans l'entête de Mesures Completes (1ère feuille: {sheet}).")
            return

        c1, c2, c3 = st.columns([1, 1, 2], vertical_alignment="center")
        with c1:
            show_surface = st.checkbox("Surfaces", value=True, key="show_surface_3d_site")
        with c2:
            show_vectors = st.checkbox("Mouvements (vecteurs)", value=False, key="show_vectors_3d_site")
        with c3:
            scale = 500.0
            if show_vectors:
                scales = _nice_scales(0, 6)
                default_val = 500 if 500 in scales else scales[0]
                if "vec_scale_3d_site" not in st.session_state:
                    st.session_state["vec_scale_3d_site"] = default_val
                scale = float(
                    st.select_slider(
                        "Échelle vecteurs",
                        options=scales,
                        value=st.session_state["vec_scale_3d_site"],
                        key="vec_scale_3d_site",
                    )
                )

        motion_minmax = (2.0, 100.0)
        if show_vectors:
            motion_minmax = st.slider(
                "Afficher seulement si mouvement (mm) entre :",
                min_value=0.0,
                max_value=2000.0,
                value=(2.0, 100.0),
                step=1.0,
                key="motion_mm_range_site",
            )

        st.markdown("#### Filtre points aberrants")
        f1, f2 = st.columns([1, 2], vertical_alignment="center")
        with f1:
            auto_filter = st.checkbox("Auto (outliers)", value=True, key="auto_filter_site")
        with f2:
            outlier_k = st.slider("Tolérance", 2.0, 20.0, 4.0, 0.5, key="outlier_k_site")

        _plot_3d(
            title="Chantier",
            selected_key="chantier_complet",
            targets=all_targets,
            mesures_path=mesures_path,
            show_surface=show_surface,
            show_vectors=show_vectors,
            scale=scale,
            auto_filter=auto_filter,
            outlier_k=float(outlier_k),
            motion_mm_range=(float(motion_minmax[0]), float(motion_minmax[1])),
            target_color_map=target_to_color,
            default_point_color="red",
        )
