from __future__ import annotations

import os
import time
from functools import lru_cache
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from src.io.json_matrix_store import read_cell as json_read_cell, write_cell as json_write_cell


# ----------------------------
# Helpers
# ----------------------------
def _abs_path(path: str) -> str:
    p = Path(path)
    if p.is_absolute():
        return str(p)
    return str((Path.cwd() / p).resolve())


def _is_cst_xlsx(abs_path: str) -> bool:
    # Detect CST workbook name
    return Path(abs_path).name.lower() == "charges sur trame.xlsx"


def _cst_runtime_folder_from_cst(abs_cst_path: str) -> Path:
    # default: beside the CST in common_data/cst_runtime
    cst = Path(abs_cst_path)
    return cst.parent / "cst_runtime"


def _should_use_json(abs_path: str) -> bool:
    p = Path(abs_path)
    if p.is_dir():
        return True
    if p.suffix.lower() == ".json":
        return True
    if p.suffix.lower() == ".xlsx" and _is_cst_xlsx(abs_path):
        rt = _cst_runtime_folder_from_cst(abs_path)
        return rt.exists() and rt.is_dir()
    return False


def _json_folder(abs_path: str) -> Path:
    p = Path(abs_path)
    if p.is_dir():
        return p
    if p.suffix.lower() == ".xlsx" and _is_cst_xlsx(abs_path):
        return _cst_runtime_folder_from_cst(abs_path)
    # If user passes a .json file, treat its parent as workbook folder (sheet name must match file name)
    if p.suffix.lower() == ".json":
        return p.parent
    return p  # fallback


# ----------------------------
# Excel cached read (pandas)
# ----------------------------
@lru_cache(maxsize=128)
def _read_sheet_xlsx(abs_path: str, sheet_name: str, mtime: float) -> pd.DataFrame:
    _ = mtime
    return pd.read_excel(abs_path, sheet_name=sheet_name, header=None, engine="openpyxl")


# ----------------------------
# JSON cached read (matrix)
# ----------------------------
@lru_cache(maxsize=256)
def _read_cell_json_cached(folder: str, sheet_name: str, row1: int, col1: int, mtime: float) -> Any:
    _ = mtime
    return json_read_cell(Path(folder), sheet_name, row1, col1)


def clear_excel_cache() -> None:
    _read_sheet_xlsx.cache_clear()
    _read_cell_json_cached.cache_clear()


# ----------------------------
# Public API (compatible avec ton projet)
# ----------------------------
def LireExcel(chemincompletclasseur, nomonglet, ligne, colonne):
    """
    Lecture cellule style VBA (1-based).
    Peut lire:
      - un .xlsx standard
      - OU un 'classeur' runtime JSON (dossier cst_runtime)
      - OU automatiquement le runtime JSON du CST si présent.
    """
    abs_path = _abs_path(str(chemincompletclasseur))

    # JSON mode
    if _should_use_json(abs_path):
        folder = _json_folder(abs_path)
        # Use sheet mtime to invalidate cache
        sheet_file = folder / f"{str(nomonglet)}.json"
        try:
            mtime = float(sheet_file.stat().st_mtime)
        except Exception:
            mtime = 0.0
        try:
            v = _read_cell_json_cached(str(folder), str(nomonglet), int(ligne), int(colonne), mtime)
            return "" if v is None else v
        except Exception:
            return ""

    # Excel mode
    try:
        mtime = float(os.path.getmtime(abs_path))
    except Exception:
        return ""

    try:
        df = _read_sheet_xlsx(abs_path, str(nomonglet), mtime)
    except Exception:
        return ""

    try:
        r = int(ligne) - 1
        c = int(colonne) - 1
    except Exception:
        return ""

    if r < 0 or c < 0 or r >= df.shape[0] or c >= df.shape[1]:
        return ""

    v = df.iat[r, c]
    if isinstance(v, float) and pd.isna(v):
        return ""
    return v


def EcrireExcel(chemincompletclasseur, nomonglet, ligne, colonne, valeur) -> None:
    """
    Écriture cellule style VBA (1-based).
    Écrit:
      - dans le JSON si runtime (ou CST runtime présent)
      - sinon dans le .xlsx via openpyxl
    """
    abs_path = _abs_path(str(chemincompletclasseur))

    # JSON mode
    if _should_use_json(abs_path):
        folder = _json_folder(abs_path)
        json_write_cell(folder, str(nomonglet), int(ligne), int(colonne), valeur)
        clear_excel_cache()
        return

    # Excel mode
    wb = load_workbook(abs_path, data_only=False, read_only=False)
    try:
        if str(nomonglet) not in wb.sheetnames:
            raise KeyError(f"Onglet introuvable: {nomonglet}")
        ws = wb[str(nomonglet)]
        ws.cell(row=int(ligne), column=int(colonne)).value = valeur
        wb.save(abs_path)
    finally:
        try:
            wb.close()
        except Exception:
            pass

    clear_excel_cache()
