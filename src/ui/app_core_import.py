from __future__ import annotations

from pathlib import Path
from datetime import datetime, timezone, date as _date
from io import BytesIO
import hashlib

import streamlit as st

from src.io.import_utils import (
    build_topo_target,
    build_inclino_target,
    build_fond_plan_target,  # ✅ AJOUT
    replace_file_bytes,
)

# ======================================================
# Formatting helpers
# ======================================================
def _fmt_bytes(n: int) -> str:
    try:
        n = int(n)
    except Exception:
        return "—"
    if n < 1024:
        return f"{n} o"
    if n < 1024**2:
        return f"{n/1024:.1f} Ko"
    if n < 1024**3:
        return f"{n/1024**2:.2f} Mo"
    return f"{n/1024**3:.2f} Go"


def _fmt_dt(dt: datetime | None) -> str:
    if not dt:
        return "—"
    return dt.astimezone().strftime("%d/%m/%Y %H:%M:%S")


def _try_get_last_modified(uploaded) -> datetime | None:
    if uploaded is None:
        return None

    for attr in ("last_modified", "lastModified", "modified", "mtime"):
        v = getattr(uploaded, attr, None)
        if v is None:
            continue

        if isinstance(v, datetime):
            return v

        try:
            fv = float(v)
            if fv > 1e12:  # ms
                return datetime.fromtimestamp(fv / 1000.0, tz=timezone.utc)
            if fv > 1e9:  # s
                return datetime.fromtimestamp(fv, tz=timezone.utc)
        except Exception:
            pass

    return None


def _content_signature(content: bytes) -> str:
    return hashlib.sha1(content).hexdigest()


# ======================================================
# Cache keys (ciblées par "dataset")
# ======================================================
def _ensure_cache_state() -> None:
    if "data_hash" not in st.session_state or not isinstance(st.session_state.get("data_hash"), dict):
        st.session_state["data_hash"] = {}
    if "data_rev" not in st.session_state or not isinstance(st.session_state.get("data_rev"), dict):
        st.session_state["data_rev"] = {}


def set_dataset_signature(kind: str, sig: str) -> None:
    _ensure_cache_state()
    st.session_state["data_hash"][kind] = sig
    st.session_state["data_rev"][kind] = int(st.session_state["data_rev"].get(kind, 0)) + 1


def get_dataset_signature(kind: str) -> str:
    _ensure_cache_state()
    return str(st.session_state["data_hash"].get(kind, ""))


def _targeted_python_cache_clear(kind: str) -> None:
    # Optional: purge de caches Python (lru_cache) si tu en as.
    try:
        if kind == "topo":
            pass
    except Exception:
        pass

    try:
        if kind == "inclino":
            pass
    except Exception:
        pass

    try:
        if kind == "fond_plan":
            pass
    except Exception:
        pass


# ======================================================
# Validation Excel (STRICT + robuste)
# ======================================================
def _parse_date_str(s: str) -> datetime | None:
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    return None


def _is_date_cell(v) -> bool:
    if v is None:
        return False
    if isinstance(v, datetime):
        return True
    if isinstance(v, _date):
        return True
    if isinstance(v, str):
        return _parse_date_str(v) is not None
    return False


def _is_number(v) -> bool:
    if v is None or isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return True
    if isinstance(v, str):
        s = v.strip().replace(",", ".")
        if not s:
            return False
        try:
            float(s)
            return True
        except Exception:
            return False
    return False


def _validate_uploaded_xlsx_strict(content: bytes) -> tuple[bool, str]:
    import openpyxl

    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as e:
        return False, f"Fichier Excel illisible: {e}"

    # 1) Une seule feuille
    if len(wb.sheetnames) != 1:
        return False, f"Le classeur doit contenir 1 seule feuille (actuel: {len(wb.sheetnames)})."

    ws = wb[wb.sheetnames[0]]

    # 2) Cibles en ligne 1 + pattern triplets (2,5,8,...)
    try:
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except Exception:
        return False, "Impossible de lire la ligne 1."

    target_cols: list[int] = []
    for idx, v in enumerate(row1, start=1):
        if idx == 1:
            continue
        if isinstance(v, str) and v.strip():
            target_cols.append(idx)

    if not target_cols:
        return False, "Aucune cible détectée en ligne 1 (colonne 2 et +)."

    bad = [c for c in target_cols if ((c - 2) % 3) != 0]
    if bad:
        return (
            False,
            "Entête invalide: les cibles doivent être en colonnes 2,5,8,... (triplets X/Y/Z). "
            f"Colonnes non conformes: {bad}",
        )

    max_col = len(row1)
    for c in target_cols:
        if c + 2 > max_col:
            return False, "Entête invalide: une cible n'a pas ses 3 colonnes (X/Y/Z) disponibles."

    # 3) Dates en colonne 1 (échantillon)
    date_checks: list[bool] = []
    non_empty = 0
    MAX_NON_EMPTY = 40

    for r in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        v = r[0] if r else None
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        non_empty += 1
        date_checks.append(_is_date_cell(v))
        if non_empty >= MAX_NON_EMPTY:
            break

    if non_empty < 8:
        return False, f"Colonne 1: pas assez de valeurs (trouvées: {non_empty}, min: 8)."

    ok_dates = sum(1 for b in date_checks if b)
    ratio = ok_dates / max(non_empty, 1)

    if ok_dates < 8 or ratio < 0.90:
        return (
            False,
            f"Colonne 1 invalide: attendu des dates (valides: {ok_dates}/{non_empty} = {ratio*100:.0f}%).",
        )

    # 4) Vérifier qu’on trouve AU MOINS un triplet XYZ numérique
    MAX_SCAN_ROWS = 800
    found_triplet = False
    scanned = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        scanned += 1
        if scanned > MAX_SCAN_ROWS:
            break

        if not row:
            continue

        d0 = row[0] if len(row) >= 1 else None
        if not _is_date_cell(d0):
            continue

        for c in target_cols[: min(len(target_cols), 24)]:
            x = row[c - 1] if (c - 1) < len(row) else None
            y = row[c] if c < len(row) else None
            z = row[c + 1] if (c + 1) < len(row) else None
            if _is_number(x) and _is_number(y) and _is_number(z):
                found_triplet = True
                break

        if found_triplet:
            break

    if not found_triplet:
        return False, f"Aucune donnée XYZ numérique détectée (scan {min(scanned, MAX_SCAN_ROWS)} lignes)."

    return True, "OK"


# ======================================================
# Validation PDF (simple mais robuste)
# ======================================================
def _validate_uploaded_pdf(content: bytes) -> tuple[bool, str]:
    if not content or len(content) < 8:
        return False, "PDF vide ou trop petit."
    if not content.startswith(b"%PDF-"):
        return False, "Signature PDF invalide (le fichier ne commence pas par %PDF-)."
    # Optionnel : limiter taille si tu veux éviter des énormes fichiers
    # if len(content) > 60 * 1024 * 1024:
    #     return False, "PDF trop volumineux (max 60 Mo)."
    return True, "OK"


# ======================================================
# Import (auto replace) + invalidation ciblée (par kind)
# ======================================================
def _auto_replace_on_upload_xlsx(kind: str, dest_path: Path, uploaded) -> None:
    if uploaded is None:
        return

    try:
        content = uploaded.getvalue()
    except Exception as e:
        st.error("❌ Impossible de lire le fichier uploadé.")
        st.exception(e)
        return

    last_mod = _try_get_last_modified(uploaded)
    st.markdown(
        f"""
- **Fichier :** `{uploaded.name}`
- **Poids :** {_fmt_bytes(getattr(uploaded, "size", len(content)))}
- **Dernière modification :** {_fmt_dt(last_mod)}
""".strip()
    )

    ok, reason = _validate_uploaded_xlsx_strict(content)
    if not ok:
        st.error(f"❌ Fichier refusé : {reason}")
        return

    _do_replace_and_invalidate(kind, dest_path, content)


def _auto_replace_on_upload_pdf(kind: str, dest_path: Path, uploaded) -> None:
    if uploaded is None:
        return

    try:
        content = uploaded.getvalue()
    except Exception as e:
        st.error("❌ Impossible de lire le fichier uploadé.")
        st.exception(e)
        return

    last_mod = _try_get_last_modified(uploaded)
    st.markdown(
        f"""
- **Fichier :** `{uploaded.name}`
- **Poids :** {_fmt_bytes(getattr(uploaded, "size", len(content)))}
- **Dernière modification :** {_fmt_dt(last_mod)}
""".strip()
    )

    ok, reason = _validate_uploaded_pdf(content)
    if not ok:
        st.error(f"❌ Fichier refusé : {reason}")
        return

    _do_replace_and_invalidate(kind, dest_path, content)


def _do_replace_and_invalidate(kind: str, dest_path: Path, content: bytes) -> None:
    sig = _content_signature(content)
    sig_key = f"import::{kind}::last_sig"
    up_key = f"import::{kind}::uploaded_at"

    prev_sig = st.session_state.get(sig_key)
    uploaded_at: datetime | None = st.session_state.get(up_key)

    if prev_sig == sig:
        st.markdown(f"- **Date d’upload :** {_fmt_dt(uploaded_at)}")
        st.success("✅ Fichier déjà importé.")
        return

    now = datetime.now(timezone.utc)

    try:
        replace_file_bytes(
            dest_path=dest_path,
            content=content,
            backup_dir=dest_path.parent / "_import_backups",
        )

        st.session_state[sig_key] = sig
        st.session_state[up_key] = now

        # ✅ invalidation ciblée
        set_dataset_signature(kind, sig)
        _targeted_python_cache_clear(kind)

        st.markdown(f"- **Date d’upload :** {_fmt_dt(now)}")
        st.success("✅ Import terminé. (Cache invalidé uniquement pour ce dataset)")
        st.rerun()

    except PermissionError as e:
        st.error(
            "❌ Permission refusée lors de l’écriture.\n\n"
            "Le fichier est probablement ouvert ou verrouillé.\n\n"
            f"Détail : {e}"
        )
    except Exception as e:
        st.error("❌ Échec de l’import.")
        st.exception(e)


def render_import(common_data_dir: Path) -> None:
    st.subheader("Import")

    tab_topo, tab_inclino, tab_fond_plan = st.tabs(["Topographie", "Inclinométrie", "Fond de plan"])  # ✅ AJOUT

    with tab_topo:
        target = build_topo_target(common_data_dir)
        st.markdown("### Topographie")
        uploaded = st.file_uploader(
            "",
            type=["xlsx"],
            key="import_topo_uploader",
            accept_multiple_files=False,
            label_visibility="collapsed",
        )
        _auto_replace_on_upload_xlsx("topo", target.dest_path, uploaded)

    with tab_inclino:
        target = build_inclino_target(common_data_dir)
        st.markdown("### Inclinométrie")
        uploaded = st.file_uploader(
            "",
            type=["xlsx"],
            key="import_inclino_uploader",
            accept_multiple_files=False,
            label_visibility="collapsed",
        )
        _auto_replace_on_upload_xlsx("inclino", target.dest_path, uploaded)

    with tab_fond_plan:
        target = build_fond_plan_target(common_data_dir)
        st.markdown("### Fond de plan")
        uploaded = st.file_uploader(
            "",
            type=["pdf"],
            key="import_fond_plan_uploader",
            accept_multiple_files=False,
            label_visibility="collapsed",
        )
        _auto_replace_on_upload_pdf("fond_plan", target.dest_path, uploaded)
