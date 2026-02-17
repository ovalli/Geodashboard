# ======================================================
# src/ui/app_core_3d_topo.py  (COMPLET - additive)
# ✅ Ne crée PAS de figure -> ajoute des traces à un fig existant
# ✅ UI (controls) inchangée: mêmes keys session_state, même comportement
# ✅ Caches/lectures identiques + cache-buster optionnel (mtime)
#
# ✅ Surface engine externalisé dans app_core_3d_topo_surface.py
# ======================================================

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import numpy as np
import streamlit as st

from src.ui.app_core_3d_topo_surface import SurfaceConfig, build_surface


# ======================================================
# Constants (conservatifs)
# ======================================================
HEAD_REL, LINE_REL, CONE_REL = 0.055, 0.015, 0.090
HEAD_ABS_MIN, LINE_ABS_MIN, CONE_ABS_MIN = 1.4, 0.5, 1.8

VECTOR_LINE_WIDTH = 6
HITBOX_SIZE = 24
POINT_SIZE = 7
_SCALE_REF = 500.0

# ======================================================
# Surface config (tu règles ici)
# ======================================================
SURF_MAX_EDGE_FACTOR = 10.0
SURF_MAX_EDGE_MIN = 10.0

SURF_ENABLE_SHARP_FILTER = True
SURF_MIN_DIHEDRAL_DEG = 45.0  # tu peux mettre 0
SURF_SHARP_TRI_DROP_MODE = "weaker"  # "weaker" ou "both"

SURF_ENABLE_QUALITY_FILTER = True
SURF_MIN_QUALITY_REL = 0.12
SURF_MIN_QUALITY_ABS = 1e-10

SURF_DROP_TINY_COMPONENTS = True
SURF_MIN_TRIANGLES_PER_COMPONENT = 5

# Excavation mask (pas encore branché UI)
SURF_ENABLE_EXCAVATION_MASK = True
SURF_EXCAVATION_MODE = "centroid"  # "centroid" ou "any_vertex"

# ======================================================
# Palette & emoji mapping (comme Coupes depuis CST)
# ======================================================
_ZONE_PALETTE = ["#146EFF", "#22C55E", "#F97316", "#A855F7", "#EF4444", "#06B6D4", "#EAB308", "#EC4899"]
_ZONE_EMOJI = ["🔵", "🟢", "🟠", "🟣", "🔴", "🟦", "🟡", "🩷"]
_COLOR_TO_EMOJI = dict(zip(_ZONE_PALETTE, _ZONE_EMOJI))


def zone_color(i: int) -> str:
    return _ZONE_PALETTE[i % len(_ZONE_PALETTE)]


def zone_emoji(i: int) -> str:
    return _ZONE_EMOJI[i % len(_ZONE_EMOJI)]


def emoji_from_color(col: str, fallback_i: int) -> str:
    return _COLOR_TO_EMOJI.get((col or "").strip(), zone_emoji(fallback_i))


# ======================================================
# Utilities
# ======================================================
def _nice_scales(min_pow: int = 0, max_pow: int = 6) -> List[int]:
    out: set[int] = set()
    for p in range(min_pow, max_pow + 1):
        b = 10**p
        out.update((1 * b, 2 * b, 5 * b))
    out.add(10 ** (max_pow + 1))
    return sorted(out)


def _fmt_date(d) -> str:
    if d is None:
        return "date inconnue"
    if hasattr(d, "strftime"):
        return d.strftime("%Y-%m-%d")
    return str(d)


def _safe_cache_buster(p) -> float:
    """
    Cache-buster robuste: combine mtime + size.
    """
    try:
        pp = p if isinstance(p, Path) else Path(str(p))
        stt = pp.stat()
        m = float(stt.st_mtime)
        s = float(getattr(stt, "st_size", 0))
        return m + (s * 1e-9)
    except Exception:
        return 0.0


# ======================================================
# Excel readers (cachés) - mtime OPTIONNEL (compat)
# ======================================================
@st.cache_data(show_spinner=False)
def mesures_header(mesures_xlsx: str, *, mtime: float = 0.0) -> Tuple[str, Dict[str, int], List[str]]:
    """
    Retourne (sheet1_name, col_map, targets) depuis la 1ère feuille.
    targets = entêtes hors colonne Date (col 1).
    """
    import openpyxl

    wb = openpyxl.load_workbook(mesures_xlsx, data_only=True, read_only=True)
    if not wb.sheetnames:
        raise ValueError("Classeur Mesures Completes sans onglets.")
    sheet = wb.sheetnames[0]
    ws = wb[sheet]

    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_map: Dict[str, int] = {
        str(v).strip(): i
        for i, v in enumerate(row1, start=1)
        if isinstance(v, str) and v.strip()
    }

    items = sorted(col_map.items(), key=lambda kv: kv[1])
    targets: list[str] = []
    for name, col in items:
        if col == 1 and name.lower() in {"date", "dates", "time", "temps"}:
            continue
        targets.append(name)

    seen: set[str] = set()
    uniq: list[str] = []
    for t in targets:
        if t not in seen:
            seen.add(t)
            uniq.append(t)

    return sheet, col_map, uniq


@st.cache_data(show_spinner=False)
def first_and_last_coords_for_targets(
    mesures_xlsx: str,
    sheet_name: str,
    col_map: Dict[str, int],
    targets: Tuple[str, ...],
    *,
    mtime: float = 0.0,
) -> Tuple[str, str, Dict[str, Tuple[float, float, float]], Dict[str, Tuple[float, float, float]]]:
    import openpyxl

    triplets = {t: (c, c + 1, c + 2) for t in targets if (c := col_map.get(t)) is not None}

    wb = openpyxl.load_workbook(mesures_xlsx, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    first_date = last_date = None
    coords_first: Dict[str, Tuple[float, float, float]] = {}
    coords_last: Dict[str, Tuple[float, float, float]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        d = row[0] if len(row) >= 1 else None
        if d is not None:
            first_date = d if first_date is None else first_date
            last_date = d

        for t, (cx, cy, cz) in triplets.items():
            if cx - 1 >= len(row) or cy - 1 >= len(row) or cz - 1 >= len(row):
                continue
            x, y, z = row[cx - 1], row[cy - 1], row[cz - 1]
            if x is None or y is None or z is None:
                continue
            try:
                pt = (float(x), float(y), float(z))
            except Exception:
                continue
            coords_first.setdefault(t, pt)
            coords_last[t] = pt

    return _fmt_date(first_date), _fmt_date(last_date), coords_first, coords_last


def _robust_keep_mask(x: np.ndarray, y: np.ndarray, z: np.ndarray, k: float) -> np.ndarray:
    pts = np.column_stack([x, y, z])
    med = np.median(pts, axis=0)
    d = np.linalg.norm(pts - med, axis=1)

    d_med = float(np.median(d))
    mad = float(np.median(np.abs(d - d_med)))
    if mad > 1e-12:
        rz = 0.6745 * (d - d_med) / mad
        return rz <= k

    q1, q3 = float(np.percentile(d, 25)), float(np.percentile(d, 75))
    iqr = q3 - q1
    if iqr <= 1e-12:
        return np.ones(len(d), dtype=bool)
    factor = max(1.5, min(6.0, k / 3.0))
    return d <= (q3 + factor * iqr)


def _render_pretty_list(items: List[str], cols: int = 3) -> None:
    items = sorted([str(x) for x in items if str(x).strip()])
    if not items:
        st.write("—")
        return
    n = len(items)
    cols = max(1, int(cols))
    rows = int(np.ceil(n / cols))
    grid = [items[i * rows : (i + 1) * rows] for i in range(cols)]
    for j, col in enumerate(st.columns(cols)):
        with col:
            for it in grid[j]:
                st.write(f"- {it}")


def _add_xyz_triad(fig, xs, ys, zs) -> None:
    import plotly.graph_objects as go

    xmin, xmax = float(xs.min()), float(xs.max())
    ymin, ymax = float(ys.min()), float(ys.max())
    zmin, zmax = float(zs.min()), float(zs.max())
    dx, dy, dz = max(xmax - xmin, 1.0), max(ymax - ymin, 1.0), max(zmax - zmin, 1.0)
    diag = float(np.sqrt(dx * dx + dy * dy + dz * dz))
    L = max(diag * 0.08, 1.0)

    ox, oy, oz = xmin + dx * 0.06, ymin + dy * 0.06, zmin + dz * 0.06

    fig.add_trace(
        go.Scatter3d(
            x=[ox, ox + L, None, ox, ox, None, ox, ox, None],
            y=[oy, oy, None, oy, oy + L, None, oy, oy, None],
            z=[oz, oz, None, oz, oz, None, oz, oz + L, None],
            mode="lines",
            hoverinfo="skip",
            line=dict(width=6, color="rgba(190,190,190,0.85)"),
            showlegend=False,
            name="",
        )
    )
    fig.add_trace(
        go.Scatter3d(
            x=[ox + L, ox, ox],
            y=[oy, oy + L, oy],
            z=[oz, oz, oz + L],
            mode="text",
            text=["X=Est", "Y=Nord", "Z"],
            hoverinfo="skip",
            textfont=dict(size=14, color="rgba(190,190,190,0.95)"),
            showlegend=False,
            name="",
        )
    )


# ======================================================
# Controls (UI)
# ======================================================
def controls(key_prefix: str, *, render_widgets: bool) -> Tuple[bool, bool, float, Tuple[float, float], bool, float]:
    k_surface = f"show_surface_3d_{key_prefix}"
    k_vectors = f"show_vectors_3d_{key_prefix}"
    k_scale = f"vec_scale_3d_{key_prefix}"
    k_motion = f"motion_mm_range_{key_prefix}"
    k_auto = f"auto_filter_{key_prefix}"
    k_k = f"outlier_k_{key_prefix}"

    st.session_state.setdefault(k_surface, True)
    st.session_state.setdefault(k_vectors, False)
    st.session_state.setdefault(k_motion, (2.0, 100.0))
    st.session_state.setdefault(k_auto, True)
    st.session_state.setdefault(k_k, 4.0)

    scales = _nice_scales(0, 6)
    st.session_state.setdefault(k_scale, 500 if 500 in scales else scales[0])

    if render_widgets:
        c1, c2, c3 = st.columns([1, 1, 2], vertical_alignment="center")
        with c1:
            st.checkbox("Surfaces", value=bool(st.session_state[k_surface]), key=k_surface)
        with c2:
            st.checkbox("Mouvements (vecteurs)", value=bool(st.session_state[k_vectors]), key=k_vectors)
        with c3:
            if bool(st.session_state[k_vectors]):
                st.select_slider("Échelle vecteurs", options=scales, value=st.session_state[k_scale], key=k_scale)

        if bool(st.session_state[k_vectors]):
            mv = st.session_state.get(k_motion, (2.0, 100.0))
            try:
                mv = (float(mv[0]), float(mv[1]))
            except Exception:
                mv = (2.0, 100.0)
            st.slider(
                "Afficher seulement si mouvement (mm) entre :",
                min_value=0.0,
                max_value=2000.0,
                value=mv,
                step=1.0,
                key=k_motion,
            )

        st.markdown("#### Filtre points aberrants")
        f1, f2 = st.columns([1, 2], vertical_alignment="center")
        with f1:
            st.checkbox("Auto (outliers)", value=bool(st.session_state[k_auto]), key=k_auto)
        with f2:
            st.slider("Tolérance", 2.0, 20.0, float(st.session_state[k_k]), 0.5, key=k_k)

    show_surface = bool(st.session_state[k_surface])
    show_vectors = bool(st.session_state[k_vectors])
    scale = float(st.session_state[k_scale]) if show_vectors else float(_SCALE_REF)

    mv = st.session_state.get(k_motion, (2.0, 100.0))
    try:
        motion = (float(mv[0]), float(mv[1]))
    except Exception:
        motion = (2.0, 100.0)

    auto_filter = bool(st.session_state[k_auto])
    outlier_k = float(st.session_state[k_k])
    return show_surface, show_vectors, scale, motion, auto_filter, outlier_k


# ======================================================
# Payload topo
# ======================================================
@dataclass
class TopoPayload:
    selected_key: str
    sheet: str
    date_first: str
    date_last: str

    names: List[str]
    xs: np.ndarray
    ys: np.ndarray
    zs: np.ndarray
    point_colors: List[str] | str

    used_scipy: bool
    tri_i: np.ndarray
    tri_j: np.ndarray
    tri_k: np.ndarray

    # debug surface
    surf_n0: int
    surf_n_edge: int
    surf_n_qual: int
    surf_n_dihedral: int
    surf_n_comp: int
    surf_n_excav: int

    vline_x: List[float]
    vline_y: List[float]
    vline_z: List[float]

    cone_x: List[float]
    cone_y: List[float]
    cone_z: List[float]
    cone_u: List[float]
    cone_v: List[float]
    cone_w: List[float]
    cone_sizeref_s: float

    caption_text: str
    warn_no_scipy: bool
    removed_outliers: List[str]
    missing: List[str]


def compute_topo_payload(
    *,
    selected_key: str,
    targets: List[str],
    mesures_path,
    show_vectors: bool,
    scale: float,
    auto_filter: bool,
    outlier_k: float,
    motion_mm_range: Tuple[float, float],
    target_color_map: Dict[str, str] | None,
    default_point_color: str,
    # (prévu pour plus tard) zones excavation en XY
    excavations_xy: Optional[List[List[Tuple[float, float]]]] = None,
) -> TopoPayload | None:
    import math

    buster = _safe_cache_buster(mesures_path)

    sheet, col_map, all_targets = mesures_header(str(mesures_path), mtime=buster)

    # ✅ FIX CRITIQUE: targets cohérents avec la source courante
    if not targets:
        eff_targets = list(all_targets)
    else:
        s_all = set(all_targets)
        eff_targets = [t for t in targets if t in s_all]
        if not eff_targets:
            eff_targets = list(all_targets)

    date_first, date_last, coords_first, coords_last = first_and_last_coords_for_targets(
        str(mesures_path), sheet, col_map, tuple(eff_targets), mtime=buster
    )

    points_last: List[Tuple[str, float, float, float]] = []
    missing: List[str] = []
    for t in eff_targets:
        pt = coords_last.get(t)
        if pt is None:
            missing.append(t)
        else:
            points_last.append((t, pt[0], pt[1], pt[2]))

    if not points_last:
        st.warning("Aucune coordonnée XYZ exploitable.")
        return None

    names_all = [p[0] for p in points_last]
    xs_all = np.array([p[1] for p in points_last], dtype=float)
    ys_all = np.array([p[2] for p in points_last], dtype=float)
    zs_all = np.array([p[3] for p in points_last], dtype=float)

    removed_outliers: List[str] = []
    if auto_filter and len(xs_all) >= 5:
        mask = _robust_keep_mask(xs_all, ys_all, zs_all, float(outlier_k))
        if mask.sum() >= 3:
            removed_outliers = [names_all[i] for i in range(len(names_all)) if not mask[i]]
            names = [names_all[i] for i in range(len(names_all)) if mask[i]]
            xs, ys, zs = xs_all[mask], ys_all[mask], zs_all[mask]
        else:
            names, xs, ys, zs = names_all, xs_all, ys_all, zs_all
    else:
        names, xs, ys, zs = names_all, xs_all, ys_all, zs_all

    xmin, xmax = float(xs.min()), float(xs.max())
    ymin, ymax = float(ys.min()), float(ys.max())
    zmin, zmax = float(zs.min()), float(zs.max())
    diag = float(math.sqrt((xmax - xmin) ** 2 + (ymax - ymin) ** 2 + (zmax - zmin) ** 2))
    diag = max(diag, 1.0)

    head_len = max(diag * HEAD_REL, HEAD_ABS_MIN)
    line_min = max(diag * LINE_REL, LINE_ABS_MIN)
    cone_sizeref = max(diag * CONE_REL, CONE_ABS_MIN)

    sf = float(_SCALE_REF) / float(scale if scale else _SCALE_REF)
    head_len_s = head_len * sf
    line_min_s = line_min * sf
    cone_sizeref_s = cone_sizeref * sf

    min_mm, max_mm = float(motion_mm_range[0]), float(motion_mm_range[1])
    min_mm = max(0.0, min_mm)
    max_mm = max(min_mm, max_mm)

    keep_set = set(names)

    # ======================================================
    # Surface build (external module)
    # ======================================================
    cfg = SurfaceConfig(
        max_edge_factor=float(SURF_MAX_EDGE_FACTOR),
        max_edge_min=float(SURF_MAX_EDGE_MIN),
        enable_quality_filter=bool(SURF_ENABLE_QUALITY_FILTER),
        min_quality_rel=float(SURF_MIN_QUALITY_REL),
        min_quality_abs=float(SURF_MIN_QUALITY_ABS),
        enable_dihedral_filter=bool(SURF_ENABLE_SHARP_FILTER) and float(SURF_MIN_DIHEDRAL_DEG) > 0.0,
        min_dihedral_deg=float(SURF_MIN_DIHEDRAL_DEG),
        dihedral_drop_mode=str(SURF_SHARP_TRI_DROP_MODE),
        drop_tiny_components=bool(SURF_DROP_TINY_COMPONENTS),
        min_tris_per_component=int(SURF_MIN_TRIANGLES_PER_COMPONENT),
        enable_excavation_mask=bool(SURF_ENABLE_EXCAVATION_MASK),
        excavation_mode=str(SURF_EXCAVATION_MODE),
    )

    tri_i, tri_j, tri_k, dbg = build_surface(xs, ys, zs, config=cfg, excavations_xy=excavations_xy)

    warn_no_scipy = False
    if (not dbg.used_scipy) and ("warned_no_scipy_delaunay" not in st.session_state):
        st.session_state["warned_no_scipy_delaunay"] = True
        warn_no_scipy = True

    # ======================================================
    # vectors
    # ======================================================
    vline_x: list[float] = []
    vline_y: list[float] = []
    vline_z: list[float] = []

    cone_x: list[float] = []
    cone_y: list[float] = []
    cone_z: list[float] = []
    cone_u: list[float] = []
    cone_v: list[float] = []
    cone_w: list[float] = []

    if show_vectors:
        for t in eff_targets:
            if t not in keep_set:
                continue
            if t not in coords_first or t not in coords_last:
                continue

            xF, yF, zF = coords_first[t]
            xL, yL, zL = coords_last[t]

            dx, dy, dz = (xL - xF), (yL - yF), (zL - zF)
            mag_m = float(math.sqrt(dx * dx + dy * dy + dz * dz))
            if mag_m <= 1e-12:
                continue

            motion_mm = mag_m * 1000.0
            if motion_mm < min_mm or motion_mm > max_mm:
                continue

            ux, uy, uz = dx / mag_m, dy / mag_m, dz / mag_m
            disp_len = mag_m / float(scale)
            total_len = max(disp_len, head_len_s + line_min_s)

            tip_x, tip_y, tip_z = xL + ux * total_len, yL + uy * total_len, zL + uz * total_len
            base_x, base_y, base_z = tip_x - ux * head_len_s, tip_y - uy * head_len_s, tip_z - uz * head_len_s

            vline_x += [xL, base_x, None]
            vline_y += [yL, base_y, None]
            vline_z += [zL, base_z, None]

            cone_x.append(base_x)
            cone_y.append(base_y)
            cone_z.append(base_z)
            cone_u.append(ux * head_len_s)
            cone_v.append(uy * head_len_s)
            cone_w.append(uz * head_len_s)

    point_colors = (
        default_point_color
        if target_color_map is None
        else [str(target_color_map.get(n, default_point_color) or default_point_color) for n in names]
    )

    caption_text = (
        f"Mesures: **{getattr(mesures_path, 'name', str(mesures_path))}** · Feuille: **{sheet}** · "
        f"Première: **{date_first}** · Dernière: **{date_last}** · "
        f"XYZ: **{len(points_last)} / {len(eff_targets)}** · "
        f"Outliers exclus: **{len(removed_outliers)}** · k={outlier_k:g}"
        + (f" · Vecteurs: mm ∈ [{min_mm:g}, {max_mm:g}]" if show_vectors else "")
        + (
            f" · Triangles: {dbg.n0}"
            + (f" → edge {dbg.n_edge}" if dbg.n_edge else "")
            + (f" → qual {dbg.n_qual}" if dbg.n_qual else "")
            + (f" → sharp {dbg.n_dihedral}" if dbg.n_dihedral else "")
            + (f" → comp {dbg.n_comp}" if dbg.n_comp else "")
            + (f" → excav {dbg.n_excav}" if excavations_xy else "")
        )
    )

    return TopoPayload(
        selected_key=selected_key,
        sheet=sheet,
        date_first=date_first,
        date_last=date_last,
        names=names,
        xs=xs,
        ys=ys,
        zs=zs,
        point_colors=point_colors,
        used_scipy=dbg.used_scipy,
        tri_i=tri_i,
        tri_j=tri_j,
        tri_k=tri_k,
        surf_n0=dbg.n0,
        surf_n_edge=dbg.n_edge,
        surf_n_qual=dbg.n_qual,
        surf_n_dihedral=dbg.n_dihedral,
        surf_n_comp=dbg.n_comp,
        surf_n_excav=dbg.n_excav,
        vline_x=vline_x,
        vline_y=vline_y,
        vline_z=vline_z,
        cone_x=cone_x,
        cone_y=cone_y,
        cone_z=cone_z,
        cone_u=cone_u,
        cone_v=cone_v,
        cone_w=cone_w,
        cone_sizeref_s=float(cone_sizeref_s),
        caption_text=caption_text,
        warn_no_scipy=warn_no_scipy,
        removed_outliers=removed_outliers,
        missing=missing,
    )


def add_topo_traces(fig, payload: TopoPayload, *, show_surface: bool, show_vectors: bool) -> None:
    import plotly.graph_objects as go

    xs, ys, zs = payload.xs, payload.ys, payload.zs
    names = payload.names

    hover_tpl = "<b>%{customdata}</b><br>Z=%{z:.1f}<extra></extra>"
    custom = np.array(names, dtype=object)

    # hitbox invisible + points visibles
    fig.add_trace(
        go.Scatter3d(
            x=xs,
            y=ys,
            z=zs,
            mode="markers",
            marker=dict(size=HITBOX_SIZE, opacity=0.0, color="rgba(255,0,0,0)"),
            customdata=custom,
            hovertemplate=hover_tpl,
            showlegend=False,
            name="",
        )
    )
    fig.add_trace(
        go.Scatter3d(
            x=xs,
            y=ys,
            z=zs,
            mode="markers",
            marker=dict(size=POINT_SIZE, opacity=1.0, color=payload.point_colors),
            customdata=custom,
            hovertemplate=hover_tpl,
            showlegend=False,
            name="",
        )
    )

    # surface
    if payload.used_scipy and len(payload.tri_i) > 0:
        fig.add_trace(
            go.Mesh3d(
                x=xs,
                y=ys,
                z=zs,
                i=payload.tri_i,
                j=payload.tri_j,
                k=payload.tri_k,
                intensity=zs,
                colorscale=[[0.0, "saddlebrown"], [1.0, "seagreen"]],
                showscale=False,
                opacity=0.95,
                flatshading=True,
                hoverinfo="skip",
                showlegend=False,
                visible=bool(show_surface and len(payload.tri_i) > 0),
                name="",
            )
        )
    else:
        fig.add_trace(
            go.Mesh3d(
                x=xs,
                y=ys,
                z=zs,
                alphahull=0,
                intensity=zs,
                colorscale=[[0.0, "saddlebrown"], [1.0, "seagreen"]],
                showscale=False,
                opacity=0.95,
                flatshading=True,
                hoverinfo="skip",
                showlegend=False,
                visible=bool(show_surface),
                name="",
            )
        )

    if show_vectors and payload.vline_x:
        fig.add_trace(
            go.Scatter3d(
                x=payload.vline_x,
                y=payload.vline_y,
                z=payload.vline_z,
                mode="lines",
                hoverinfo="skip",
                line=dict(color="orange", width=int(VECTOR_LINE_WIDTH)),
                showlegend=False,
                name="",
            )
        )

    if show_vectors and payload.cone_x:
        fig.add_trace(
            go.Cone(
                x=payload.cone_x,
                y=payload.cone_y,
                z=payload.cone_z,
                u=payload.cone_u,
                v=payload.cone_v,
                w=payload.cone_w,
                anchor="tail",
                cauto=False,
                cmin=0.0,
                cmax=1.0,
                colorscale=[[0.0, "orange"], [1.0, "orange"]],
                showscale=False,
                sizemode="absolute",
                sizeref=float(payload.cone_sizeref_s),
                opacity=0.98,
                hoverinfo="skip",
                name="",
                lighting=dict(ambient=1.0, diffuse=0.0, specular=0.0, roughness=1.0, fresnel=0.0),
            )
        )

    _add_xyz_triad(fig, xs, ys, zs)


def topo_post_render(payload: TopoPayload, *, show_surface: bool) -> None:
    st.caption(payload.caption_text)

    if show_surface and payload.warn_no_scipy:
        st.warning(
            "⚠️ SciPy non disponible : surface calculée en mode fallback (qualité moindre). "
            "Installe `scipy` pour retrouver la surface propre."
        )

    if payload.removed_outliers:
        with st.expander(f"Cibles exclues (outliers) ({len(payload.removed_outliers)})"):
            _render_pretty_list(payload.removed_outliers, cols=3)

    if payload.missing:
        with st.expander(f"Sans XYZ trouvés ({len(payload.missing)})"):
            _render_pretty_list(payload.missing, cols=3)
