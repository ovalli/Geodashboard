from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd


# ============================================================
# Helpers datetime (même logique que courbes_plotly)
# ============================================================

def _to_datetime_series(s: pd.Series) -> pd.Series:
    if s.empty:
        return pd.to_datetime(s)
    if pd.api.types.is_datetime64_any_dtype(s):
        return s
    if pd.api.types.is_numeric_dtype(s):
        # Excel serial dates
        return pd.to_datetime("1899-12-30") + pd.to_timedelta(s, unit="D")
    return pd.to_datetime(s, errors="coerce")


def _is_blank(v) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and pd.isna(v):
        return True
    return str(v).strip() == ""


def _resolve_path(workbook_path: str) -> str:
    return str(Path(workbook_path).expanduser().resolve())


def _file_mtime(workbook_path_resolved: str) -> float:
    try:
        return Path(workbook_path_resolved).stat().st_mtime
    except Exception:
        return 0.0


def _file_mtime_plus_size(workbook_path_resolved: str) -> float:
    """Cache-buster robuste: combine mtime + size (comme app_core_3d_topo)."""
    try:
        p = Path(workbook_path_resolved)
        stt = p.stat()
        return float(stt.st_mtime) + float(getattr(stt, "st_size", 0)) * 1e-9
    except Exception:
        return _file_mtime(workbook_path_resolved)


@dataclass(frozen=True)
class TargetBlock:
    name: str
    col_x: int
    col_y: int
    col_z: int


def _infer_target_blocks(df: pd.DataFrame) -> List[TargetBlock]:
    """
    Structure attendue :
      - Col 0 : Date
      - Puis blocs de 3 colonnes : X, Y, Z
      - Nom de la cible en ligne 0, colonne X (début du bloc)
        ex: row0[col_x] = "PC101" puis row0[col_x+1] vide, row0[col_x+2] vide (souvent)
    """
    if df.shape[0] < 2 or df.shape[1] < 4:
        return []

    header = df.iloc[0, :]
    blocks: List[TargetBlock] = []

    c = 1
    while c + 2 < df.shape[1]:
        v = header.iloc[c]
        if not _is_blank(v):
            name = str(v).strip()
            blocks.append(TargetBlock(name=name, col_x=c, col_y=c + 1, col_z=c + 2))
            c += 3
        else:
            c += 1

    # dédoublonnage conservateur (au cas où)
    seen = set()
    out: List[TargetBlock] = []
    for b in blocks:
        key = b.name.strip()
        if key and key not in seen:
            out.append(b)
            seen.add(key)
    return out


@lru_cache(maxsize=32)
def _read_workbook_sheet_cached(workbook_path_resolved: str, sheet_name: Optional[str], mtime: float) -> pd.DataFrame:
    """
    Lit la feuille (ou la 1ère feuille si sheet_name None).
    Important: header=None car structure custom.

    ✅ mtime dans la signature => si le fichier change, le cache est invalidé.
    """
    if sheet_name:
        return pd.read_excel(workbook_path_resolved, sheet_name=sheet_name, header=None, engine="openpyxl")

    # 1ère feuille
    xls = pd.ExcelFile(workbook_path_resolved, engine="openpyxl")
    first = xls.sheet_names[0]
    return pd.read_excel(workbook_path_resolved, sheet_name=first, header=None, engine="openpyxl")


def _read_workbook_sheet(
    workbook_path: str,
    sheet_name: Optional[str],
    mtime: Optional[float],
) -> pd.DataFrame:
    wp = _resolve_path(workbook_path)
    mt = float(_file_mtime(wp) if mtime is None else mtime)
    return _read_workbook_sheet_cached(wp, sheet_name, mt)


@lru_cache(maxsize=32)
def _list_targets_cached(workbook_path_resolved: str, sheet_name: Optional[str], mtime: float) -> List[str]:
    df = _read_workbook_sheet_cached(workbook_path_resolved, sheet_name, mtime)
    blocks = _infer_target_blocks(df)
    return [b.name for b in blocks]


def list_targets_in_mesures(
    workbook_path: str,
    sheet_name: Optional[str] = None,
    mtime: Optional[float] = None,
) -> List[str]:
    """
    ✅ mtime optionnel : si fourni, il sert de clé de cache.
    """
    wp = _resolve_path(workbook_path)
    mt = float(_file_mtime(wp) if mtime is None else mtime)
    return _list_targets_cached(wp, sheet_name, mt)


def read_targets_timeseries(
    workbook_path: str,
    targets: Iterable[str],
    sheet_name: Optional[str] = None,
    mtime: Optional[float] = None,
) -> Dict[str, pd.DataFrame]:
    """
    Retourne {target_name: df} avec colonnes: date, x, y, z (valeurs brutes).

    ✅ mtime optionnel : force l’invalidation du cache de lecture du classeur.
    """
    # ------------------------------------------------------------
    # FAST PATH: lecture streaming openpyxl (read_only) + colonnes ciblées
    # → beaucoup plus rapide que pandas.read_excel() sur gros classeur.
    # ------------------------------------------------------------
    wp = _resolve_path(workbook_path)
    mt = float(_file_mtime_plus_size(wp) if mtime is None else mtime)
    tgt = tuple([str(t).strip() for t in targets if str(t).strip()])
    if not tgt:
        return {}

    try:
        return _read_targets_timeseries_openpyxl_cached(wp, sheet_name, tgt, mt)
    except Exception:
        # Fallback conservateur: méthode pandas historique (compat / debug)
        df = _read_workbook_sheet(workbook_path, sheet_name, mtime)
        blocks = _infer_target_blocks(df)
        if not blocks:
            return {}

        block_by_name = {b.name: b for b in blocks}

        out: Dict[str, pd.DataFrame] = {}
        for name in targets:
            b = block_by_name.get(name)
            if b is None:
                continue

            dates = _to_datetime_series(df.iloc[1:, 0])
            x = pd.to_numeric(df.iloc[1:, b.col_x], errors="coerce")
            y = pd.to_numeric(df.iloc[1:, b.col_y], errors="coerce")
            z = pd.to_numeric(df.iloc[1:, b.col_z], errors="coerce")

            d = pd.DataFrame({"date": dates, "x": x, "y": y, "z": z})
            d = d.dropna(subset=["date"]).sort_values("date").reset_index(drop=True)

            # On garde les lignes où au moins une coord existe
            d = d.dropna(subset=["x", "y", "z"], how="all")

            if not d.empty:
                out[name] = d

        return out


# ============================================================
# FAST openpyxl streaming reader
# ============================================================

@lru_cache(maxsize=64)
def _header_triplets_cached(
    workbook_path_resolved: str,
    sheet_name: Optional[str],
    mtime: float,
) -> Tuple[str, Dict[str, Tuple[int, int, int]]]:
    """Retourne (sheet_used, {target: (cx,cy,cz)}) en indices 0-based."""
    import openpyxl

    wb = openpyxl.load_workbook(workbook_path_resolved, data_only=True, read_only=True)
    if not wb.sheetnames:
        raise ValueError("Classeur Mesures Completes sans onglets.")
    sheet_used = sheet_name if (sheet_name and sheet_name in wb.sheetnames) else wb.sheetnames[0]
    ws = wb[sheet_used]

    row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

    triplets: Dict[str, Tuple[int, int, int]] = {}
    # Structure: col 0 = Date, puis blocs: X,Y,Z; nom en début de bloc
    for i, v in enumerate(row1):
        if i == 0:
            continue
        if isinstance(v, str) and v.strip():
            name = v.strip()
            # blocs de 3 colonnes à partir de i
            triplets.setdefault(name, (i, i + 1, i + 2))

    return sheet_used, triplets


@lru_cache(maxsize=32)
def _read_targets_timeseries_openpyxl_cached(
    workbook_path_resolved: str,
    sheet_name: Optional[str],
    targets: Tuple[str, ...],
    mtime: float,
) -> Dict[str, pd.DataFrame]:
    import openpyxl

    sheet_used, triplets_all = _header_triplets_cached(workbook_path_resolved, sheet_name, mtime)
    triplets = {t: triplets_all[t] for t in targets if t in triplets_all}
    if not triplets:
        return {}

    wb = openpyxl.load_workbook(workbook_path_resolved, data_only=True, read_only=True)
    ws = wb[sheet_used]

    # Buffers
    dates: list = []
    xs: Dict[str, list] = {t: [] for t in triplets}
    ys: Dict[str, list] = {t: [] for t in triplets}
    zs: Dict[str, list] = {t: [] for t in triplets}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        d = row[0] if len(row) >= 1 else None
        if d is None:
            continue

        # Pour rester compatible pandas: on ajoute une ligne si au moins UNE cible a une coord
        any_val = False
        row_vals: Dict[str, Tuple[object, object, object]] = {}
        for t, (cx, cy, cz) in triplets.items():
            x = row[cx] if cx < len(row) else None
            y = row[cy] if cy < len(row) else None
            z = row[cz] if cz < len(row) else None
            if x is not None or y is not None or z is not None:
                any_val = True
            row_vals[t] = (x, y, z)

        if not any_val:
            continue

        dates.append(d)
        for t, (x, y, z) in row_vals.items():
            xs[t].append(x)
            ys[t].append(y)
            zs[t].append(z)

    if not dates:
        return {}

    # Conversion date -> datetime (même logique que _to_datetime_series)
    s_dates = pd.Series(dates)
    s_dates = _to_datetime_series(s_dates)

    out: Dict[str, pd.DataFrame] = {}
    for t in targets:
        if t not in triplets:
            continue
        x = pd.to_numeric(pd.Series(xs[t]), errors="coerce")
        y = pd.to_numeric(pd.Series(ys[t]), errors="coerce")
        z = pd.to_numeric(pd.Series(zs[t]), errors="coerce")

        ddf = pd.DataFrame({"date": s_dates, "x": x, "y": y, "z": z})
        ddf = ddf.dropna(subset=["date"]).sort_values("date").reset_index(drop=True)
        ddf = ddf.dropna(subset=["x", "y", "z"], how="all")
        if not ddf.empty:
            out[t] = ddf

    return out


def compute_deltas_vs_first_known(df: pd.DataFrame) -> Optional[pd.DataFrame]:
    """
    Prend un df date,x,y,z et retourne date, dx,dy,dz
    delta vs première mesure connue (où x,y,z tous présents).
    """
    if df is None or df.empty:
        return None

    base = df.dropna(subset=["x", "y", "z"], how="any")
    if base.empty:
        return None

    x0 = float(base.iloc[0]["x"])
    y0 = float(base.iloc[0]["y"])
    z0 = float(base.iloc[0]["z"])

    d = df.copy()
    d["dx"] = pd.to_numeric(d["x"], errors="coerce") - x0
    d["dy"] = pd.to_numeric(d["y"], errors="coerce") - y0
    d["dz"] = pd.to_numeric(d["z"], errors="coerce") - z0

    d = d[["date", "dx", "dy", "dz"]].dropna(subset=["date"])
    return d
