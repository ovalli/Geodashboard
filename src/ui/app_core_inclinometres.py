from __future__ import annotations

from pathlib import Path
import os
import streamlit as st
from openpyxl import load_workbook
from streamlit_option_menu import option_menu


def _find_inclino_excel(common_data_folder: Path) -> Path | None:
    """
    Retourne le 1er fichier .xlsx dans common_data dont le nom contient 'inclino'
    (insensible à la casse), en ignorant les fichiers temporaires Excel.
    """
    if not common_data_folder.exists():
        return None

    candidates = sorted(
        p for p in common_data_folder.iterdir()
        if p.is_file()
        and p.suffix.lower() == ".xlsx"
        and not p.name.startswith("~$")
        and "inclino" in p.name.lower()
    )
    return candidates[0] if candidates else None


@st.cache_data(show_spinner=False)
def _get_sheetnames_cached(xlsx_path: str, mtime: float) -> list[str]:
    """
    Lit les noms de feuilles (cache invalidé dès que le fichier change).
    """
    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    return list(wb.sheetnames)


def render_inclinometres(common_data_folder: Path) -> None:
    """
    Page Inclinomètres:
    - trouve le fichier common_data/*inclino*.xlsx
    - affiche un menu horizontal avec 1 item par feuille
    - contenu vide pour l'instant
    """
    xlsx_path = _find_inclino_excel(common_data_folder)
    if xlsx_path is None:
        st.warning("Aucun fichier 'inclino' trouvé dans common_data.")
        return

    mtime = os.path.getmtime(xlsx_path)
    sheetnames = _get_sheetnames_cached(str(xlsx_path), mtime)

    if not sheetnames:
        st.warning("Le fichier inclino ne contient aucune feuille.")
        return

    # Menu horizontal (style identique à Paramétrage si tu utilises déjà option_menu)
    selected_sheet = option_menu(
        menu_title=None,
        options=sheetnames,
        icons=None,
        orientation="horizontal",
        default_index=0,
        styles={
            # si tu veux EXACTEMENT comme Paramétrage, tu peux reprendre les mêmes styles
            # sinon laisse ce bloc tel quel (c'est sobre et stable)
            "container": {"padding": "0!important", "margin": "0!important"},
            "nav-link": {"padding": "10px 14px", "font-size": "14px"},
            "nav-link-selected": {"font-weight": "600"},
        },
    )

    # Pour le moment: rien dedans (placeholder)
    st.markdown(f"#### {selected_sheet}")
    st.info("Contenu à définir plus tard.")
