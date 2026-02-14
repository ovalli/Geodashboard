from __future__ import annotations

from pathlib import Path
from typing import Any

import os
import re
import tempfile

import pandas as pd
import streamlit as st
from openpyxl import load_workbook


# ======================================================
# Helpers
# ======================================================
def _as_path(p: str | Path) -> Path:
    return p if isinstance(p, Path) else Path(str(p))


def _key(session_prefix: str, sheet_name: str, suffix: str) -> str:
    if session_prefix:
        return f"{session_prefix}::{sheet_name}::{suffix}"
    return f"{sheet_name}::{suffix}"


def _stringify_col(c: Any) -> str:
    if c is None:
        return ""
    s = str(c)
    return s.strip()


def _dedupe_columns(cols: list[Any]) -> list[str]:
    """
    Streamlit data_editor exige des noms de colonnes uniques.
    - Convertit tout en str
    - Remplace vides par _col_XX
    - Dédoublonne en ajoutant __2, __3, etc.
    """
    base = [_stringify_col(c) for c in cols]

    # 1) Remplace "Unnamed: ..." et vides par des noms uniques
    out: list[str] = []
    unnamed_count = 0
    for s in base:
        if not s or s.lower().startswith("unnamed:"):
            unnamed_count += 1
            out.append(f"_col_{unnamed_count:02d}")
        else:
            out.append(s)

    # 2) Dédoublonnage général
    seen: dict[str, int] = {}
    final: list[str] = []
    for s in out:
        if s not in seen:
            seen[s] = 1
            final.append(s)
        else:
            seen[s] += 1
            final.append(f"{s}__{seen[s]}")

    return final


def _load_sheet_df(xlsx_path: Path, sheet_name: str) -> pd.DataFrame:
    # dtype=object => conserve mieux les strings / mixes
    df = pd.read_excel(
        xlsx_path,
        sheet_name=sheet_name,
        engine="openpyxl",
        dtype=object,
    )

    if df is None:
        df = pd.DataFrame()

    # ✅ IMPORTANT: rendre les colonnes uniques pour st.data_editor
    df.columns = _dedupe_columns(list(df.columns))

    return df


def _apply_freeze_panes(ws, freeze_first_row: bool, freeze_first_column: bool) -> None:
    if not freeze_first_row and not freeze_first_column:
        ws.freeze_panes = None
        return

    row = 2 if freeze_first_row else 1
    col = 2 if freeze_first_column else 1
    ws.freeze_panes = ws.cell(row=row, column=col).coordinate


def _write_df_to_sheet(ws, df: pd.DataFrame) -> None:
    """
    Écrit df en A1, avec headers.
    On écrase les valeurs dans la zone nécessaire (sans toucher aux styles autant que possible).

    Note: si on a renommé des colonnes vides (_col_01...), on ré-écrit ces noms.
    Si tu veux garder header vide dans Excel, on peut faire une règle inverse, mais
    ça re-crée le problème côté Streamlit. Là on privilégie robustesse.
    """
    if df is None:
        df = pd.DataFrame()

    df2 = df.copy()
    df2 = df2.where(pd.notnull(df2), None)

    nrows = int(df2.shape[0])
    ncols = int(df2.shape[1])

    old_max_row = int(ws.max_row or 1)
    old_max_col = int(ws.max_column or 1)

    target_rows = max(old_max_row, nrows + 1)  # +1 header
    target_cols = max(old_max_col, ncols)

    # clear values in target zone
    for r in range(1, target_rows + 1):
        for c in range(1, target_cols + 1):
            ws.cell(row=r, column=c).value = None

    # header
    for c_idx, col_name in enumerate(df2.columns, start=1):
        ws.cell(row=1, column=c_idx).value = col_name

    # data
    for r_idx in range(nrows):
        for c_idx in range(ncols):
            ws.cell(row=r_idx + 2, column=c_idx + 1).value = df2.iat[r_idx, c_idx]


def _save_workbook_atomic(wb, dst_path: Path) -> None:
    """
    Sauvegarde robuste : write tmp puis os.replace.
    """
    tmp_dir = str(dst_path.parent)
    fd, tmp_path = tempfile.mkstemp(
        prefix=dst_path.stem + "_",
        suffix=".tmp.xlsx",
        dir=tmp_dir,
    )
    os.close(fd)

    try:
        wb.save(tmp_path)
        try:
            with open(tmp_path, "rb") as f:
                os.fsync(f.fileno())
        except Exception:
            pass
        os.replace(tmp_path, str(dst_path))
    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass


# ======================================================
# Public API (COMPAT)
# ======================================================
def render_excel_sheet_editor(*args, **kwargs) -> None:
    """
    Éditeur Excel Streamlit (compatible avec plusieurs signatures historiques).

    Accepte notamment :
      - render_excel_sheet_editor(xlsx_path, sheet_name, ...)
      - render_excel_sheet_editor(workbook_path=..., sheet_name=..., ...)
      - render_excel_sheet_editor(xlsx_abs_path=..., sheet="Paramètres", ...)
      - + session_prefix=... (évite collisions de keys)

    Params supportés (keywords):
      xlsx_path, workbook_path, xlsx_abs_path, file_path
      sheet_name, sheet
      title
      freeze_first_row, freeze_first_column
      session_prefix
    """
    title: str | None = kwargs.pop("title", None)
    session_prefix: str = kwargs.pop("session_prefix", "")

    freeze_first_row: bool = bool(kwargs.pop("freeze_first_row", True))
    freeze_first_column: bool = bool(kwargs.pop("freeze_first_column", False))

    xlsx_any = (
        kwargs.pop("xlsx_path", None)
        or kwargs.pop("workbook_path", None)
        or kwargs.pop("xlsx_abs_path", None)
        or kwargs.pop("file_path", None)
        or kwargs.pop("path", None)
    )
    sheet_name = kwargs.pop("sheet_name", None) or kwargs.pop("sheet", None)

    # support appel positionnel: (xlsx_path, sheet_name, ...)
    if len(args) >= 1 and xlsx_any is None:
        xlsx_any = args[0]
    if len(args) >= 2 and sheet_name is None:
        sheet_name = args[1]

    # ignore le reste (évite crash sur legacy kwargs)
    _unused = kwargs  # noqa

    if title:
        st.subheader(title)

    if not xlsx_any:
        st.error("render_excel_sheet_editor: chemin du fichier Excel manquant (xlsx_path/workbook_path/...).")
        return
    if not sheet_name or not str(sheet_name).strip():
        st.error("render_excel_sheet_editor: nom de feuille manquant (sheet_name/sheet).")
        return

    xlsx = _as_path(xlsx_any)
    sheet_name = str(sheet_name).strip()

    if not xlsx.exists():
        st.error(f"Fichier Excel introuvable : {xlsx}")
        return

    # ---- load dataframe ----
    try:
        df = _load_sheet_df(xlsx, sheet_name)
    except ValueError as e:
        st.error(f"Feuille introuvable : {sheet_name}\n\nDétail: {e}")
        return
    except Exception as e:
        st.error(f"Lecture Excel impossible : {type(e).__name__}: {e}")
        return

    k_editor = _key(session_prefix, sheet_name, "editor")
    k_save = _key(session_prefix, sheet_name, "save")
    k_reload = _key(session_prefix, sheet_name, "reload")

    colA, colB = st.columns([1, 1])
    with colA:
        save_clicked = st.button("💾 Enregistrer", key=k_save, use_container_width=True)
    with colB:
        reload_clicked = st.button("↻ Recharger", key=k_reload, use_container_width=True)

    if reload_clicked:
        try:
            if k_editor in st.session_state:
                del st.session_state[k_editor]
        except Exception:
            pass
        st.rerun()

    edited = st.data_editor(
        df,
        key=k_editor,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
    )

    if not save_clicked:
        return

    # ---- save ----
    try:
        wb = load_workbook(str(xlsx), data_only=False, read_only=False)
        try:
            if sheet_name not in wb.sheetnames:
                st.error(f"Feuille introuvable au moment de sauvegarder : {sheet_name}")
                return

            ws = wb[sheet_name]
            _write_df_to_sheet(ws, edited)
            _apply_freeze_panes(ws, freeze_first_row=freeze_first_row, freeze_first_column=freeze_first_column)

            _save_workbook_atomic(wb, xlsx)

        finally:
            try:
                wb.close()
            except Exception:
                pass

        st.success(f"✅ Enregistré : {xlsx.name} / {sheet_name}")
        st.rerun()

    except PermissionError as e:
        st.error(
            "❌ Impossible d’écrire dans le fichier (PermissionError).\n\n"
            "Cause la plus fréquente : le fichier est **ouvert dans Excel** (verrouillé).\n\n"
            f"Détail: {e}"
        )
    except Exception as e:
        st.error(f"❌ Sauvegarde échouée : {type(e).__name__}: {e}")
